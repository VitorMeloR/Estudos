from openpyxl import Workbook, load_workbook

def matriz_sod():
    try:
        wb = load_workbook('matriz.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws['A2'] = '1°sistema'
        ws['A3'] = '2°sistema'
        wb.save('matriz.xlsx')
    coluna = ws.max_column+1
    while True:
        conflito1 = input('Diga o primeiro sistema conflitante: ').strip().upper()
        conflito2 = input('Diga o segundo sistema conflitante: ').strip().upper()
        cadastro = False
        for coluna_atual in ws.iter_cols(min_row=2, max_row=3, min_col= coluna-1, max_col= coluna, values_only=True):
            if (conflito1, conflito2) == coluna_atual or (conflito2, conflito1) == coluna_atual:
                cadastro = True
                print('Sistemas conflitantes já adicionados')
                break

        if cadastro == False:
            ws.cell(row=1, column=coluna, value=f'conflitante_{coluna - 1}')
            ws.cell(row=2, column=coluna, value=conflito1)
            ws.cell(row=3, column=coluna, value=conflito2)
            print('Sistenas conflitantes adicionados com sucesso')

        esc = input('Deseja adiconar um novo conlfito? [s/n]').strip().lower()[0]
        if esc in 'Nn':
            break
    wb.save('matriz.xlsx')
    print('sistema encerrado')

