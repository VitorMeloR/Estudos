
from openpyxl import Workbook, load_workbook
import string
# Função para coletar e validar o login do usuário
def login(login_in):
    while True:
        login = str(login_in).strip()
        if len(login) <= 15 and len(login) >= 5:
            break
        else:
           return False
    return login

# Função para coletar e validar a senha do usuário
def senha(senha_in):
    especias = list(string.punctuation)
    while True:
        senha = str(senha_in).strip()
        if len(senha) > 8:
            has_special = any(char in senha for char in especias)
            has_digit = any(char.isdigit() for char in senha)
            has_upper = any(char.isupper() for char in senha)

            if has_special and has_digit and has_upper:
                break
            else:
                return False
        else:
            return 8

    return senha

# Função para coletar e formatar o CPF do usuário
def cpf(cpf_in):
    while True:
        cpf = str(cpf_in).strip()
        cpf_l = cpf.replace('.','').replace('-','').replace(',','')
        if len(cpf_l) == 11:
            cpf1 = cpf_l[0:3]
            cpf2 = cpf_l[3:6]
            cpf3 = cpf_l[6:9]
            cpf4 = cpf_l[9:12]
            cpf_formatado = f'{cpf1}.{cpf2}.{cpf3}-{cpf4}'
            return cpf_formatado
            break
        else:
            return False

# Função principal para coletar dados e registrar no arquivo Excel
def cadastrar_dados(login,senha,cpf):
    try:
        wb = load_workbook('dados.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'LOGIN'
        ws['B1'] = 'SENHA'
        ws['C1'] = 'CPF'

    linha = ws.max_row + 1
    esc = ''

    while True:
        login_input = login
        senha_input = senha
        cpf_input = cpf
        cadastrado = False

        # Verifica se o login ou CPF já existem no arquivo Excel
        for linha_atual in ws.iter_rows(min_row=2, values_only=True):
            if login_input == linha_atual[0] or cpf_input == linha_atual[2]:
                cadastrado = True
                return False
                break

        # Se não estiver cadastrado, insere os dados no arquivo Excel
        if not cadastrado:
            ws.cell(row=linha, column=1, value=login_input)
            ws.cell(row=linha, column=2, value=senha_input)
            ws.cell(row=linha, column=3, value=cpf_input)
            wb.save('dados.xlsx')
            break
    wb.save('dados.xlsx')

# Se este arquivo for executado diretamente, chama a função principal
if __name__ == "__main__":
    cadastrar_dados()

