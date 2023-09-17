from openpyxl import Workbook, load_workbook
import requests

try:
    wb = load_workbook('data_base.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = 'Sistemas'
    ws['A1'] = 'Cod_sistemas'
    ws['B1'] = 'Sistemas'
def cadastrar_sistema():
    cadastrado = False
    linha = ws.max_row + 1
    cod_sistema = f'SO{ws.max_row}'
    sistema = requests.get('http://127.0.0.1:5000')
    print(sistema)

    for linha_atual in ws.iter_rows(min_row=2, values_only=True):
        if sistema == linha_atual[1] or cod_sistema == linha_atual[0]:
            cadastrado = True
            break

    # Se não estiver cadastrado, insere os dados no arquivo Excel
    if not cadastrado:
        ws.cell(row=linha, column=1, value=cod_sistema)
        ws.cell(row=linha, column=2, value=sistema)
        wb.save('data_base.xlsx')
        print('Sistema cadastrado com sucesso!')
    else:
        print('O sistema já está cadastrado.')

    wb.close()

# Call the function to run your code
cadastrar_sistema()
