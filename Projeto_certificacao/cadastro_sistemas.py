from openpyxl import Workbook, load_workbook

try:
    wb = load_workbook('dados.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = 'Sistemas'
    ws['A1'] = 'Cod_sistemas'
    ws['B1'] = 'Sistemas'
def cadastrar_sistema():
    cadastrado = False
    linha = ws.max_row + 1
    cod_sistema = f'SO{ws.max_row}'
    sistema = input('Digite o nome do sistema: ')

    for linha_atual in ws.iter_rows(min_row=2, values_only=True):
        if sistema == linha_atual[1] or cod_sistema == linha_atual[0]:
            cadastrado = True
            break

    # Se não estiver cadastrado, insere os dados no arquivo Excel
    if not cadastrado:
        ws.cell(row=linha, column=1, value=cod_sistema)
        ws.cell(row=linha, column=2, value=sistema)
        wb.save('dados.xlsx')
        print('Sistema cadastrado com sucesso!')
    else:
        print('O sistema já está cadastrado.')

    wb.close()

# Call the function to run your code
cadastrar_sistema()
